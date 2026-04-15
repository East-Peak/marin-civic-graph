#!/usr/bin/env python3
"""Normalize NetFile campaign finance ZIP exports into settled-ontology JSONL for Neo4j loading.

Reads NetFile Excel workbooks (ZIP → .xlsx) and produces:
  - nodes.jsonl  (Committee, MoneyFlow, Person, Organization, Place nodes)
  - edges.jsonl  (FROM_SOURCE, TO_TARGET, EVIDENCED_BY, IN_JURISDICTION)
  - normalization-report.json

Usage:
  python scripts/normalize_campaign_finance.py --source marin-county-campaign-finance
  python scripts/normalize_campaign_finance.py --all
  python scripts/normalize_campaign_finance.py --source marin-county-campaign-finance --load
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import zipfile
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))

ROOT = Path(__file__).resolve().parent.parent


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def slugify_name(last: str | None, first: str | None = None) -> str:
    """Produce a stable hyphen-slug from a name pair."""
    parts = []
    if last:
        parts.append(last.strip())
    if first:
        parts.append(first.strip())
    raw = " ".join(parts)
    # Strip possessive / punctuation-only characters before lowercasing,
    # so "O'Brien" → "OBrien" (no extra hyphen) rather than "O-Brien".
    slug = re.sub(r"[''`]", "", raw)
    slug = slug.lower()
    slug = re.sub(r"[^a-z0-9]+", "-", slug)
    slug = re.sub(r"-{2,}", "-", slug)
    slug = slug.strip("-")
    return slug


def _date_str(val) -> str | None:
    """Convert openpyxl date cell value (datetime or string) to ISO YYYY-MM-DD, or None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str) and val.strip():
        return val.strip()
    return None


def _node(
    id: str,
    node_type: str,
    labels: list[str],
    display_label: str,
    properties: dict,
    capture_id: str,
    section: str = "campaign_finance",
    status: str | None = None,
) -> dict:
    return {
        "id": id,
        "node_type": node_type,
        "labels": labels,
        "display_label": display_label,
        "promotion_state": "promoted",
        "source_bundle_ids": [capture_id],
        "source_sections": [section],
        "source_status": status,
        "properties": properties,
        "qa_lane": False,
    }


def _edge(
    source_id: str,
    source_type: str,
    target_id: str,
    target_type: str,
    rel_type: str,
    capture_id: str,
    properties: dict | None = None,
) -> dict:
    return {
        "source_id": source_id,
        "source_node_type": source_type,
        "target_id": target_id,
        "target_node_type": target_type,
        "relationship_type": rel_type,
        "source_bundle_ids": [capture_id],
        "source_fields": ["normalize_campaign_finance"],
        "properties": properties or {},
    }


# ---------------------------------------------------------------------------
# Node builders
# ---------------------------------------------------------------------------

def build_committee_node(
    filer_id: int | str,
    filer_name: str,
    committee_type: str,
    jurisdiction_id: str,
    capture_id: str,
) -> dict:
    # filer_id is usually a numeric string but can be "Pending" for newly-registered committees
    try:
        filer_id_stored: int | str = int(filer_id)
    except (TypeError, ValueError):
        filer_id_stored = str(filer_id) if filer_id is not None else ""
    node_id = f"committee-netfile-{filer_id_stored}"
    return _node(
        id=node_id,
        node_type="Committee",
        labels=["Committee"],
        display_label=filer_name or node_id,
        properties={
            "name": filer_name,
            "netfile_filer_id": filer_id_stored,
            "committee_type": committee_type,
            "jurisdiction_id": jurisdiction_id,
        },
        capture_id=capture_id,
        section="committee_stubs",
        status="stub_from_netfile_export",
    )


def build_moneyflow_node(
    filer_id: int | str,
    tran_id: str,
    amount: float,
    flow_date: str | None,
    flow_type: str,
    source_schedule: str,
    capture_id: str,
) -> dict:
    node_id = f"moneyflow-{filer_id}-{tran_id}"
    props: dict = {
        "amount": amount,
        "flow_type": flow_type,
        "source_schedule": source_schedule,
    }
    if flow_date:
        props["flow_date"] = flow_date
    return _node(
        id=node_id,
        node_type="MoneyFlow",
        labels=["MoneyFlow"],
        display_label=f"{flow_type} ${amount:.2f}",
        properties=props,
        capture_id=capture_id,
        section="money_flows",
        status="from_netfile_export",
    )


def build_contributor_node(
    name_last: str | None,
    name_first: str | None,
    entity_cd: str,
    capture_id: str,
) -> dict:
    slug = slugify_name(name_last, name_first)
    org_entity_codes = {"COM", "OTH", "SCC", "PTY"}
    if entity_cd in org_entity_codes:
        node_id = f"org-{slug}"
        display = name_last or slug
        return _node(
            id=node_id,
            node_type="Organization",
            labels=["Organization"],
            display_label=display,
            properties={"name": display, "entity_cd": entity_cd},
            capture_id=capture_id,
            section="contributor_stubs",
            status="stub_from_netfile_export",
        )
    else:
        # IND (individual) and anything else → Person
        node_id = f"person-{slug}"
        parts = [p for p in [name_first, name_last] if p and p.strip()]
        display = " ".join(parts) if parts else slug
        return _node(
            id=node_id,
            node_type="Person",
            labels=["Person"],
            display_label=display,
            properties={"name": display, "entity_cd": entity_cd},
            capture_id=capture_id,
            section="contributor_stubs",
            status="stub_from_netfile_export",
        )


# ---------------------------------------------------------------------------
# Excel parsers
# ---------------------------------------------------------------------------

def _sheet_headers(ws) -> dict[str, int]:
    """Return column-name → 0-based index mapping from row 1."""
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return {str(v): i for i, v in enumerate(first_row) if v is not None}


def _cell(row: tuple, headers: dict[str, int], col: str):
    idx = headers.get(col)
    if idx is None:
        return None
    return row[idx] if idx < len(row) else None


def parse_contributions(zip_path: Path) -> list[dict]:
    """Parse A-Contributions sheet from a NetFile ZIP export."""
    rows = []
    with zipfile.ZipFile(zip_path) as zf:
        xlsx_name = next(n for n in zf.namelist() if n.endswith(".xlsx"))
        with zf.open(xlsx_name) as f:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb["A-Contributions"]
            headers = _sheet_headers(ws)
            for raw in ws.iter_rows(min_row=2, values_only=True):
                amount = _cell(raw, headers, "Tran_Amt1")
                if not amount:
                    continue
                try:
                    amount = float(amount)
                except (TypeError, ValueError):
                    continue
                if amount == 0:
                    continue
                rows.append({
                    "filer_id": _cell(raw, headers, "Filer_ID"),
                    "filer_name": _cell(raw, headers, "Filer_NamL"),
                    "committee_type": _cell(raw, headers, "Committee_Type"),
                    "tran_id": _cell(raw, headers, "Tran_ID"),
                    "entity_cd": _cell(raw, headers, "Entity_Cd"),
                    "contributor_last": _cell(raw, headers, "Tran_NamL"),
                    "contributor_first": _cell(raw, headers, "Tran_NamF"),
                    "amount": amount,
                    "flow_date": _date_str(_cell(raw, headers, "Tran_Date")),
                    "employer": _cell(raw, headers, "Tran_Emp"),
                    "occupation": _cell(raw, headers, "Tran_Occ"),
                    "city": _cell(raw, headers, "Tran_City"),
                    "state": _cell(raw, headers, "Tran_State"),
                    "zip": _cell(raw, headers, "Tran_Zip4"),
                    "elect_date": _date_str(_cell(raw, headers, "Elect_Date")),
                })
    return rows


def parse_expenditures(zip_path: Path) -> list[dict]:
    """Parse E-Expenditure sheet from a NetFile ZIP export."""
    rows = []
    with zipfile.ZipFile(zip_path) as zf:
        xlsx_name = next(n for n in zf.namelist() if n.endswith(".xlsx"))
        with zf.open(xlsx_name) as f:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb["E-Expenditure"]
            headers = _sheet_headers(ws)
            for raw in ws.iter_rows(min_row=2, values_only=True):
                amount = _cell(raw, headers, "Amount")
                if not amount:
                    continue
                try:
                    amount = float(amount)
                except (TypeError, ValueError):
                    continue
                if amount == 0:
                    continue
                rows.append({
                    "filer_id": _cell(raw, headers, "Filer_ID"),
                    "filer_name": _cell(raw, headers, "Filer_NamL"),
                    "committee_type": _cell(raw, headers, "Committee_Type"),
                    "tran_id": _cell(raw, headers, "Tran_ID"),
                    "entity_cd": _cell(raw, headers, "Entity_Cd"),
                    "payee_last": _cell(raw, headers, "Payee_NamL"),
                    "payee_first": _cell(raw, headers, "Payee_NamF"),
                    "amount": amount,
                    "flow_date": _date_str(_cell(raw, headers, "Expn_Date")),
                    "city": _cell(raw, headers, "Payee_City"),
                    "state": _cell(raw, headers, "Payee_State"),
                    "zip": _cell(raw, headers, "Payee_Zip4"),
                    "elect_date": _date_str(_cell(raw, headers, "Elect_Date")),
                    "expn_code": _cell(raw, headers, "Expn_Code"),
                    "expn_dscr": _cell(raw, headers, "Expn_Dscr"),
                })
    return rows


# ---------------------------------------------------------------------------
# Source normalizer
# ---------------------------------------------------------------------------

def normalize_campaign_source(
    capture: dict,
    zip_paths: list[Path],
    output_dir: Path,
) -> tuple[list[dict], list[dict], dict]:
    """Parse ZIP exports and write settled-format JSONL + report."""
    output_dir.mkdir(parents=True, exist_ok=True)

    capture_id = capture["capture_id"]
    jurisdiction_id = capture["jurisdiction_id"]
    source_id = capture["source_id"]

    nodes: list[dict] = []
    edges: list[dict] = []

    # Dedup registries keyed by slug / filer_id
    committees: dict[str, dict] = {}      # filer_id str → node
    contributors: dict[str, dict] = {}    # slug → node

    # Place stub
    place_node = _node(
        id=jurisdiction_id,
        node_type="Place",
        labels=["Place"],
        display_label=jurisdiction_id.replace("place-", "").replace("-", " ").title(),
        properties={"name": jurisdiction_id.replace("place-", "").replace("-", " ").title()},
        capture_id=capture_id,
        section="place_stubs",
        status="stub_from_source_config",
    )
    nodes.append(place_node)

    moneyflow_count = 0

    for zip_path in zip_paths:
        year = zip_path.stem  # e.g. "2024"
        record_id = f"record-{source_id}-export-{year}"

        # --- Contributions ---
        contrib_rows = parse_contributions(zip_path)
        for row in contrib_rows:
            filer_id = str(row["filer_id"])

            # Committee node (dedup)
            if filer_id not in committees:
                committee_node = build_committee_node(
                    filer_id=filer_id,
                    filer_name=row["filer_name"],
                    committee_type=row["committee_type"] or "",
                    jurisdiction_id=jurisdiction_id,
                    capture_id=capture_id,
                )
                committees[filer_id] = committee_node

            # Contributor node (dedup by slug)
            entity_cd = row["entity_cd"] or "IND"
            contributor_slug = slugify_name(row["contributor_last"], row["contributor_first"])
            if contributor_slug and contributor_slug not in contributors:
                contributor_node = build_contributor_node(
                    name_last=row["contributor_last"],
                    name_first=row["contributor_first"],
                    entity_cd=entity_cd,
                    capture_id=capture_id,
                )
                contributors[contributor_slug] = contributor_node

            # MoneyFlow node
            moneyflow_node = build_moneyflow_node(
                filer_id=filer_id,
                tran_id=row["tran_id"],
                amount=row["amount"],
                flow_date=row["flow_date"],
                flow_type="contribution",
                source_schedule="A",
                capture_id=capture_id,
            )
            nodes.append(moneyflow_node)
            moneyflow_count += 1
            mf_id = moneyflow_node["id"]
            committee_id = f"committee-netfile-{filer_id}"

            # Edges
            if contributor_slug:
                contributor_id = contributors[contributor_slug]["id"]
                contributor_type = contributors[contributor_slug]["node_type"]
                # contributor → moneyflow → committee
                edges.append(_edge(contributor_id, contributor_type, mf_id, "MoneyFlow",
                                   "FROM_SOURCE", capture_id))
            edges.append(_edge(mf_id, "MoneyFlow", committee_id, "Committee",
                               "TO_TARGET", capture_id))
            edges.append(_edge(mf_id, "MoneyFlow", record_id, "Record",
                               "EVIDENCED_BY", capture_id))

        # --- Expenditures ---
        expend_rows = parse_expenditures(zip_path)
        for row in expend_rows:
            filer_id = str(row["filer_id"])

            # Committee node (dedup)
            if filer_id not in committees:
                committee_node = build_committee_node(
                    filer_id=filer_id,
                    filer_name=row["filer_name"],
                    committee_type=row["committee_type"] or "",
                    jurisdiction_id=jurisdiction_id,
                    capture_id=capture_id,
                )
                committees[filer_id] = committee_node

            # Payee node (dedup by slug)
            entity_cd = row["entity_cd"] or "OTH"
            payee_slug = slugify_name(row["payee_last"], row["payee_first"])
            if payee_slug and payee_slug not in contributors:
                payee_node = build_contributor_node(
                    name_last=row["payee_last"],
                    name_first=row["payee_first"],
                    entity_cd=entity_cd,
                    capture_id=capture_id,
                )
                contributors[payee_slug] = payee_node

            # MoneyFlow node
            moneyflow_node = build_moneyflow_node(
                filer_id=filer_id,
                tran_id=row["tran_id"],
                amount=row["amount"],
                flow_date=row["flow_date"],
                flow_type="expenditure",
                source_schedule="E",
                capture_id=capture_id,
            )
            nodes.append(moneyflow_node)
            moneyflow_count += 1
            mf_id = moneyflow_node["id"]
            committee_id = f"committee-netfile-{filer_id}"

            # Edges: committee → moneyflow → payee
            edges.append(_edge(committee_id, "Committee", mf_id, "MoneyFlow",
                               "FROM_SOURCE", capture_id))
            if payee_slug:
                payee_id = contributors[payee_slug]["id"]
                payee_type = contributors[payee_slug]["node_type"]
                edges.append(_edge(mf_id, "MoneyFlow", payee_id, payee_type,
                                   "TO_TARGET", capture_id))
            edges.append(_edge(mf_id, "MoneyFlow", record_id, "Record",
                               "EVIDENCED_BY", capture_id))

    # Add deduped committees and contributors to node list
    for committee_node in committees.values():
        nodes.append(committee_node)
        # Committee → Place edge
        edges.append(_edge(committee_node["id"], "Committee", jurisdiction_id, "Place",
                           "IN_JURISDICTION", capture_id))

    for contributor_node in contributors.values():
        nodes.append(contributor_node)

    # Write JSONL
    with open(output_dir / "nodes.jsonl", "w") as f:
        for node in nodes:
            f.write(json.dumps(node, sort_keys=True) + "\n")
    with open(output_dir / "edges.jsonl", "w") as f:
        for edge in edges:
            f.write(json.dumps(edge, sort_keys=True) + "\n")

    report = {
        "source_id": source_id,
        "capture_id": capture_id,
        "normalized_at": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "node_count": len(nodes),
        "edge_count": len(edges),
        "committee_count": len(committees),
        "contributor_count": len(contributors),
        "moneyflow_count": moneyflow_count,
    }
    with open(output_dir / "normalization-report.json", "w") as f:
        json.dump(report, indent=2, fp=f)
        f.write("\n")

    return nodes, edges, report


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _find_zip_paths(source_id: str) -> list[Path]:
    """Discover all yearly ZIP exports under data/raw/{source_id}/*/"""
    pattern = ROOT / "data" / "raw" / source_id / "*" / "*.zip"
    return sorted(pattern.parent.parent.glob("*/*.zip"))


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Normalize NetFile campaign finance exports to settled-ontology JSONL"
    )
    parser.add_argument("--source", help="Source ID to normalize")
    parser.add_argument("--all", dest="all_sources", action="store_true",
                        help="Normalize all NetFile sources")
    parser.add_argument("--load", action="store_true",
                        help="Load into Neo4j after normalization")
    args = parser.parse_args()

    import yaml

    sources = []
    registry_path = ROOT / "registry" / "netfile-sources.yaml"
    if registry_path.exists():
        with open(registry_path) as f:
            data = yaml.safe_load(f)
        sources.extend(data.get("sources", []))

    if args.source:
        targets = [s for s in sources if s["id"] == args.source]
        if not targets:
            print(f"Unknown source: {args.source}", file=sys.stderr)
            sys.exit(1)
    elif args.all_sources:
        targets = sources
    else:
        print("Specify --source <id> or --all", file=sys.stderr)
        sys.exit(1)

    for source_config in targets:
        source_id = source_config["id"]
        zip_paths = _find_zip_paths(source_id)
        if not zip_paths:
            print(f"  No ZIP exports found for {source_id}, skipping")
            continue

        # Use the capture date from the most recent capture dir
        capture_date = sorted((ROOT / "data" / "raw" / source_id).iterdir())[-1].name
        capture_id = f"{source_id}__{capture_date}"
        capture = {
            "source_id": source_id,
            "capture_id": capture_id,
            "jurisdiction_id": source_config["jurisdiction_id"],
            "institution_id": source_config["institution_id"],
            "captured_at": f"{capture_date}T00:00:00Z",
        }

        print(f"\nNormalizing: {source_id}")
        print(f"  ZIPs: {[p.name for p in zip_paths]}")

        output_dir = ROOT / "data" / "normalized" / f"{source_id}-campaign-finance"
        nodes, edges, report = normalize_campaign_source(capture, zip_paths, output_dir)

        print(f"  Committees:   {report['committee_count']}")
        print(f"  Contributors: {report['contributor_count']}")
        print(f"  MoneyFlows:   {report['moneyflow_count']}")
        print(f"  Nodes total:  {report['node_count']}")
        print(f"  Edges total:  {report['edge_count']}")
        print(f"  Output:       {output_dir}")

        if args.load:
            from load_neo4j_v2 import load_nodes as neo4j_load_nodes, load_edges as neo4j_load_edges

            from neo4j import GraphDatabase

            uri = os.getenv("NEO4J_URI")
            user = os.getenv("NEO4J_USER")
            password = os.getenv("NEO4J_PASSWORD")
            if not all([uri, user, password]):
                print("  NEO4J_URI, NEO4J_USER, NEO4J_PASSWORD required for --load",
                      file=sys.stderr)
                continue

            driver = GraphDatabase.driver(uri, auth=(user, password))
            try:
                print("  Loading into Neo4j...")
                neo4j_load_nodes(driver, nodes)
                neo4j_load_edges(driver, edges)
                print("  Loaded.")
            finally:
                driver.close()


if __name__ == "__main__":
    main()
